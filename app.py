import streamlit as st
from openpyxl import load_workbook
import pandas as pd
from jinja2 import Template
import io

st.set_page_config(page_title="Excel إلى HTML", layout="wide")
st.title("📊 تحويل ملف Excel إلى صفحة HTML تفاعلية")

uploaded_file = st.file_uploader("ارفع ملف Excel", type=["xlsx"])

if uploaded_file:
    in_memory_file = io.BytesIO(uploaded_file.read())
    wb = load_workbook(in_memory_file, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data = []

    for row in ws.iter_rows(min_row=2):
        if all(cell.value in [None, ""] for cell in row):
            continue
        row_data = {}
        for i, cell in enumerate(row):
            if i >= 16:
                continue  # Ignore columns beyond 16
            header = headers[i] if i < len(headers) else f"col{i}"
            if header in ["الموقع", "صور"] and cell.hyperlink:
                label = "📍 الموقع" if header == "الموقع" else "📷 صور"
                value = f'<a href="{cell.hyperlink.target}" target="_blank" style="padding:5px 10px;background-color:#6c757d;color:white;border-radius:5px;text-decoration:none;">{label}</a>'
            else:
                value = cell.value if cell.value is not None else ""
            row_data[header] = value
        if len(row_data) == 16:
            data.append(row_data)

    df = pd.DataFrame(data)

    # Ensure exactly 16 columns
    df = df.iloc[:, :16]

    # Convert to HTML table body
    table_rows = df.to_html(index=False, header=False, escape=False).split("<tbody>")[1].split("</tbody>")[0]

    with open("template_corrected.html", "r", encoding="utf-8") as f:
        template = Template(f.read())

    html_result = template.render(table_rows=table_rows)

    st.download_button("📥 تحميل HTML", html_result, file_name="شقق_للإيجار.html", mime="text/html")